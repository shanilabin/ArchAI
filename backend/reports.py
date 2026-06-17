from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


def generate_quantity_survey_excel(project_name: str, data: list):
    """
    מקבל רשימת שורות (כל שורה dict עם המפתחות:
    'Item Code', 'Description', 'Unit', 'Quantity', 'Unit Price')
    ומחזיר קובץ Excel מעוצב (bytes). מחושב גם עמודת Total.
    כתוב עם openpyxl בלבד — בלי pandas.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quantity Survey"

    # שורת כותרת
    headers = ["Item Code", "Description", "Unit", "Quantity", "Unit Price", "Total"]
    ws.append(headers)

    # שורות נתונים + חישוב Total
    for row in data:
        quantity = row.get("Quantity", 0) or 0
        unit_price = row.get("Unit Price", 0) or 0
        total = quantity * unit_price
        ws.append([
            row.get("Item Code", ""),
            row.get("Description", ""),
            row.get("Unit", ""),
            quantity,
            unit_price,
            total,
        ])

    # --- עיצוב ---
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # כותרת
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # נתונים
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            cell.alignment = center
            cell.border = border
            # פורמט מספרי לעמודות Quantity / Unit Price / Total
            if cell.column_letter in ["D", "E", "F"]:
                cell.number_format = "#,##0.00"

    # רוחב עמודות אוטומטי
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
